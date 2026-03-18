"""Utilities for importing product data from PDF, XLSX, CSV, and ODS files.

Supports the specific formats used by Winaris Glass LLC:
- BENSON LIST style: Part Number | PRODUCT | COLOR | VEHICLE
- Invoice style:     Item | Description | Price | Quantity On Hand
- Generic column mapping for any spreadsheet
"""
import re
import csv
from io import BytesIO, StringIO
from datetime import datetime

import pdfplumber
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_col(header, keywords):
    for i, h in enumerate(header):
        for kw in keywords:
            if kw in h:
                return i
    return None


def _safe_cell(row, idx):
    if idx is not None and idx < len(row) and row[idx]:
        return str(row[idx]).strip()
    return ""


def _safe_int(row, idx, default=0):
    try:
        return int(float(_safe_cell(row, idx)))
    except (ValueError, TypeError):
        return default


def _safe_float(row, idx, default=0.0):
    try:
        return float(_safe_cell(row, idx).replace(",", "").replace("$", ""))
    except (ValueError, TypeError):
        return default


def _to_float(val):
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _to_int(val):
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def _parse_vehicle_string(vehicle_str):
    """Parse vehicle strings like 'CHEVROLET SILVERADO PICKUP 07-14' or 'CAVALIER 4D 2003'."""
    if not vehicle_str:
        return "", "", None, None

    vehicle_str = vehicle_str.strip()

    # Try to find year range at end: '07-14', '2020-', '2016-', '2003'
    year_match = re.search(r'(\d{2,4})\s*[-–]\s*(\d{2,4})?\s*$', vehicle_str)
    year_start = None
    year_end = None
    base = vehicle_str

    if year_match:
        base = vehicle_str[:year_match.start()].strip()
        y1 = year_match.group(1)
        y2 = year_match.group(2)
        if len(y1) == 2:
            y1_int = int(y1)
            year_start = 2000 + y1_int if y1_int < 50 else 1900 + y1_int
        else:
            year_start = int(y1)
        if y2:
            if len(y2) == 2:
                y2_int = int(y2)
                year_end = 2000 + y2_int if y2_int < 50 else 1900 + y2_int
            else:
                year_end = int(y2)
        else:
            year_end = datetime.now().year
    else:
        single_year = re.search(r'\b((?:19|20)\d{2})\b\s*$', vehicle_str)
        if single_year:
            base = vehicle_str[:single_year.start()].strip()
            year_start = int(single_year.group(1))
            year_end = year_start

    # Split base into make and model
    # Remove common suffixes that aren't make/model
    base = re.sub(r'\b(PICKUP|SEDAN|COUPE|SUV|VAN|WAGON|4D|2D|5D|CAB|BACK WINDOW)\b', ' ', base, flags=re.IGNORECASE)
    base = re.sub(r'\s+', ' ', base).strip()

    parts = base.split(None, 1)
    car_make = parts[0].title() if parts else ""
    car_model = parts[1].title() if len(parts) > 1 else ""

    return car_make, car_model, year_start, year_end


# ---------------------------------------------------------------------------
# Unified row reader — reads rows from XLSX, CSV, or ODS into a list of lists
# ---------------------------------------------------------------------------

def _read_rows_from_file(file_storage, filename):
    """Read rows from any supported file format. Returns list of lists."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    file_bytes = file_storage.read()
    file_storage.seek(0)

    if ext == "csv":
        return _read_csv_rows(file_bytes)
    elif ext == "ods":
        return _read_ods_rows(file_bytes)
    elif ext in ("xlsx", "xls"):
        return _read_xlsx_rows(file_bytes)
    else:
        raise ValueError(f"Unsupported file format: .{ext}")


def _read_xlsx_rows(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_csv_rows(file_bytes):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(StringIO(text))
    return [row for row in reader]


def _read_ods_rows(file_bytes):
    from odf.opendocument import load as ods_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = ods_load(BytesIO(file_bytes))
    rows = []
    for table in doc.getElementsByType(Table):
        for tr in table.getElementsByType(TableRow):
            row = []
            for cell in tr.getElementsByType(TableCell):
                repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                texts = cell.getElementsByType(P)
                val = " ".join(t.firstChild.data if t.firstChild else "" for t in texts).strip()
                # Try to convert to number
                val_attr = cell.getAttribute("value")
                if val_attr:
                    try:
                        val = float(val_attr)
                        if val == int(val):
                            val = int(val)
                    except (ValueError, TypeError):
                        pass
                for _ in range(min(repeat, 20)):
                    row.append(val if val != "" else None)
            rows.append(row)
        break  # Only read the first table/sheet
    return rows


def _find_header_row(rows):
    """Find the first row that looks like a header (has text in multiple cells)."""
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:
            return i
    return 0


# ---------------------------------------------------------------------------
# Smart product parser — detects file format and parses accordingly
# ---------------------------------------------------------------------------

def parse_products_file(file_storage, filename):
    """
    Parse any supported file (XLSX, CSV, ODS) and return a list of product dicts.
    Auto-detects the column structure:
    - BENSON LIST style: Part Number | PRODUCT | COLOR | VEHICLE
    - Invoice style:     Item | Description | Price | Quantity On Hand
    - Generic column mapping
    """
    rows = _read_rows_from_file(file_storage, filename)
    if len(rows) < 2:
        return []

    header_idx = _find_header_row(rows)
    header = [str(c).lower().strip() if c else "" for c in rows[header_idx]]

    # Detect format
    if _is_benson_format(header):
        return _parse_benson_rows(rows[header_idx + 1:], header)
    elif _is_invoice_format(header):
        return _parse_invoice_rows(rows[header_idx + 1:], header)
    else:
        return _parse_generic_rows(rows[header_idx + 1:], header)


def _is_benson_format(header):
    """Detect: Part Number | PRODUCT | COLOR | VEHICLE"""
    h = " ".join(header)
    return ("part number" in h or "part no" in h) and "vehicle" in h


def _is_invoice_format(header):
    """Detect: Item | Description | Price | Quantity On Hand"""
    h = " ".join(header)
    return "item" in h and "price" in h and "quantity" in h


def _parse_benson_rows(rows, header):
    """Parse BENSON LIST format: Part Number | PRODUCT | COLOR | VEHICLE"""
    pn_idx = _find_col(header, ["part number", "part no", "part #"])
    prod_idx = _find_col(header, ["product", "item"])
    color_idx = _find_col(header, ["color", "colour"])
    vehicle_idx = _find_col(header, ["vehicle", "car", "auto"])

    products = []
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        part_full = _safe_cell(row, pn_idx)
        product_code = _safe_cell(row, prod_idx)
        color = _safe_cell(row, color_idx)
        vehicle = _safe_cell(row, vehicle_idx)

        if not part_full and not product_code:
            continue

        part_number = product_code or part_full.split()[0] if part_full else ""
        car_make, car_model, year_start, year_end = _parse_vehicle_string(vehicle)

        name = vehicle if vehicle else part_number
        if color:
            desc = f"Part: {part_number}, Color: {color}"
        else:
            desc = f"Part: {part_number}"

        products.append({
            "name": name,
            "part_number": part_number,
            "description": desc,
            "car_make": car_make,
            "car_model": car_model,
            "car_year_start": year_start,
            "car_year_end": year_end,
            "category": _guess_category(name, vehicle),
            "price": 0.0,
            "cost": 0.0,
            "stock_quantity": 0,
        })
    return products


def _parse_invoice_rows(rows, header):
    """Parse invoice format: Item | Description | Price | Quantity On Hand"""
    item_idx = _find_col(header, ["item", "part", "sku", "code"])
    desc_idx = _find_col(header, ["description", "desc", "name"])
    price_idx = _find_col(header, ["price", "cost", "amount"])
    qty_idx = _find_col(header, ["quantity", "qty", "stock", "on hand"])

    products = []
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        part_number = _safe_cell(row, item_idx)
        description = _safe_cell(row, desc_idx)
        price = _safe_float(row, price_idx, 0.0)
        qty = _safe_int(row, qty_idx, 0)

        if not part_number and not description:
            continue

        car_make, car_model, year_start, year_end = _parse_vehicle_string(description)
        name = description if description else part_number

        products.append({
            "name": name,
            "part_number": part_number,
            "description": description,
            "car_make": car_make,
            "car_model": car_model,
            "car_year_start": year_start,
            "car_year_end": year_end,
            "category": _guess_category(name, description),
            "price": price,
            "cost": price,
            "stock_quantity": qty,
        })
    return products


GENERIC_COLUMNS = {
    "name": ["name", "product", "description", "item", "part name", "glass type"],
    "category": ["category", "type", "glass category"],
    "car_make": ["make", "car make", "vehicle make", "manufacturer"],
    "car_model": ["model", "car model", "vehicle model"],
    "car_year_start": ["year start", "year from", "from year", "start year"],
    "car_year_end": ["year end", "year to", "to year", "end year"],
    "part_number": ["part number", "part no", "part #", "sku", "code", "oem number"],
    "price": ["price", "sell price", "selling price", "retail price"],
    "cost": ["cost", "unit cost", "buy price", "wholesale price"],
    "stock_quantity": ["stock", "quantity", "qty", "inventory", "stock quantity"],
    "description": ["description", "notes", "details", "info"],
}


def _parse_generic_rows(rows, header):
    """Parse generic spreadsheet with flexible column mapping."""
    col_map = {}
    for field, keywords in GENERIC_COLUMNS.items():
        for i, h in enumerate(header):
            for kw in keywords:
                if kw in h:
                    col_map[field] = i
                    break
            if field in col_map:
                break

    products = []
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        product = {}
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            product[field] = val if val is not None else ""

        if not product.get("name"):
            continue

        product["price"] = _to_float(product.get("price", 0))
        product["cost"] = _to_float(product.get("cost", 0))
        product["stock_quantity"] = _to_int(product.get("stock_quantity", 0))
        product["car_year_start"] = _to_int(product.get("car_year_start"))
        product["car_year_end"] = _to_int(product.get("car_year_end"))
        product.setdefault("category", "General")
        product.setdefault("description", "")
        product.setdefault("part_number", "")
        product.setdefault("car_make", "")
        product.setdefault("car_model", "")

        products.append(product)
    return products


def _guess_category(name, description=""):
    """Guess product category from name/description text."""
    text = (name + " " + (description or "")).lower()
    if any(w in text for w in ["windshield", "front glass", "back glass", "rear glass", "back window"]):
        return "Windshield"
    if any(w in text for w in ["side window", "door glass", "front window", "left", "right"]):
        return "Side Window"
    if any(w in text for w in ["quarter", "vent"]):
        return "Quarter Glass"
    if any(w in text for w in ["sunroof", "moonroof", "panoramic"]):
        return "Sunroof"
    if any(w in text for w in ["mirror"]):
        return "Mirror"
    return "Auto Glass"


# ---------------------------------------------------------------------------
# PDF Invoice Parsing
# ---------------------------------------------------------------------------

def parse_pdf_invoice(file_storage):
    raw_text = ""
    tables = []

    pdf_bytes = file_storage.read()
    file_storage.seek(0)

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            raw_text += page_text + "\n"
            page_tables = page.extract_tables()
            if page_tables:
                tables.extend(page_tables)

    result = {
        "invoice_number": _extract_invoice_number(raw_text),
        "supplier": _extract_supplier(raw_text),
        "date": _extract_date(raw_text),
        "raw_text": raw_text.strip(),
        "items": [],
        "total": 0.0,
    }

    if tables:
        result["items"] = _parse_pdf_table_items(tables)
    else:
        result["items"] = _parse_pdf_text_items(raw_text)

    result["total"] = sum(
        (item.get("quantity", 0) * item.get("unit_cost", 0)) for item in result["items"]
    )
    return result


def _extract_invoice_number(text):
    for pat in [r"(?:Invoice|Inv)[\s#.:]*(\S+)", r"(?:Invoice Number|Invoice No)[:\s]*(\S+)", r"#\s*(\d{3,})"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return f"PDF-{datetime.now().strftime('%Y%m%d%H%M%S')}"


def _extract_supplier(text):
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    return lines[0][:120] if lines else "Unknown Supplier"


def _extract_date(text):
    for pat in [r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", r"(\d{4}[/-]\d{1,2}[/-]\d{1,2})", r"(?:Date)[:\s]*(\S+)"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1)
    return datetime.now().strftime("%Y-%m-%d")


def _parse_pdf_table_items(tables):
    items = []
    for table in tables:
        if not table or len(table) < 2:
            continue
        header = [str(c).lower().strip() if c else "" for c in table[0]]
        name_idx = _find_col(header, ["description", "item", "product", "name", "part"])
        qty_idx = _find_col(header, ["qty", "quantity", "count"])
        cost_idx = _find_col(header, ["price", "cost", "unit", "rate", "amount"])
        part_idx = _find_col(header, ["part no", "part #", "part number", "sku", "code"])
        for row in table[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            item = {
                "name": _safe_cell(row, name_idx),
                "quantity": _safe_int(row, qty_idx, 1),
                "unit_cost": _safe_float(row, cost_idx, 0.0),
                "part_number": _safe_cell(row, part_idx),
            }
            if item["name"]:
                items.append(item)
    return items


def _parse_pdf_text_items(text):
    items = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"(.+?)\s+(\d+)\s+\$?([\d,]+\.?\d*)", line)
        if m:
            items.append({"name": m.group(1).strip(), "quantity": int(m.group(2)),
                          "unit_cost": float(m.group(3).replace(",", "")), "part_number": ""})
    return items


# ---------------------------------------------------------------------------
# Export — supports XLSX, CSV, ODS
# ---------------------------------------------------------------------------

def _product_rows(products):
    """Convert product objects to header + data rows."""
    headers = ["Part Number", "Name", "Category", "Car Make", "Car Model",
               "Year Start", "Year End", "Price", "Cost", "Stock Qty", "Description"]
    rows = [headers]
    for p in products:
        rows.append([
            p.part_number, p.name, p.category, p.car_make, p.car_model,
            p.car_year_start, p.car_year_end, p.price, p.cost, p.stock_quantity,
            p.description,
        ])
    return rows


def _order_rows(orders):
    headers = ["Order ID", "Customer", "Email", "Phone", "Address",
               "Status", "Total", "Items", "Date", "Notes"]
    rows = [headers]
    for o in orders:
        item_summary = "; ".join(f"{it.product.name} x{it.quantity}" for it in o.items if it.product)
        rows.append([
            o.id, o.customer_name, o.customer_email, o.customer_phone,
            o.customer_address, o.status, o.total_amount, item_summary,
            o.created_at.strftime("%Y-%m-%d") if o.created_at else "", o.notes,
        ])
    return rows


def _invoice_rows(invoices):
    headers = ["ID", "Invoice #", "Supplier", "Date", "Total", "Status", "Notes"]
    rows = [headers]
    for inv in invoices:
        rows.append([
            inv.id, inv.invoice_number, inv.supplier,
            inv.date.strftime("%Y-%m-%d") if inv.date else "",
            inv.total_amount, inv.status, inv.notes,
        ])
    return rows


def _export_xlsx(rows, sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _export_csv(rows):
    buf = StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow([str(c) if c is not None else "" for c in row])
    result = BytesIO(buf.getvalue().encode("utf-8-sig"))
    result.seek(0)
    return result


def _export_ods(rows, sheet_name="Sheet1"):
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    table = Table(name=sheet_name)
    for row_data in rows:
        tr = TableRow()
        for val in row_data:
            cell = TableCell()
            cell.addElement(P(text=str(val) if val is not None else ""))
            tr.addElement(cell)
        table.addElement(tr)
    doc.spreadsheet.addElement(table)
    buf = BytesIO()
    doc.save(buf)
    buf.seek(0)
    return buf


def export_products(products, fmt="xlsx"):
    rows = _product_rows(products)
    if fmt == "csv":
        return _export_csv(rows), "products.csv", "text/csv"
    elif fmt == "ods":
        return _export_ods(rows, "Products"), "products.ods", "application/vnd.oasis.opendocument.spreadsheet"
    else:
        return _export_xlsx(rows, "Products"), "products.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_orders(orders, fmt="xlsx"):
    rows = _order_rows(orders)
    if fmt == "csv":
        return _export_csv(rows), "orders.csv", "text/csv"
    elif fmt == "ods":
        return _export_ods(rows, "Orders"), "orders.ods", "application/vnd.oasis.opendocument.spreadsheet"
    else:
        return _export_xlsx(rows, "Orders"), "orders.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_invoices(invoices, fmt="xlsx"):
    rows = _invoice_rows(invoices)
    if fmt == "csv":
        return _export_csv(rows), "invoices.csv", "text/csv"
    elif fmt == "ods":
        return _export_ods(rows, "Invoices"), "invoices.ods", "application/vnd.oasis.opendocument.spreadsheet"
    else:
        return _export_xlsx(rows, "Invoices"), "invoices.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# Keep old names for backward compat
def parse_excel_products(file_storage):
    return parse_products_file(file_storage, "upload.xlsx")

def export_products_to_excel(products):
    buf, _, _ = export_products(products, "xlsx")
    return buf

def export_orders_to_excel(orders):
    buf, _, _ = export_orders(orders, "xlsx")
    return buf

def export_invoices_to_excel(invoices):
    buf, _, _ = export_invoices(invoices, "xlsx")
    return buf
